import datetime
from collections import defaultdict
from decimal import Decimal

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import Faculty, Room, ScheduleConfig, ScheduleEntry


def _format_time(t):
    """Format a time object as '8:00 AM' style, portable across platforms."""
    hour = t.hour % 12 or 12
    minute = f'{t.minute:02d}'
    ampm = 'AM' if t.hour < 12 else 'PM'
    return f'{hour}:{minute} {ampm}'


def export_schedule(tenant, period):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Schedule'

    headers = [
        'Course', 'In-Charge', '', 'Course\nCode', 'Course Title',
        'Lec\nUnits', 'Lab\nUnits', 'Course\nUnits', 'Contact\nHours',
        '', '', 'Faculty', 'Faculty\nCredits', 'Day(s)', 'Time In',
        'Time Out', 'Room', 'Section', 'Load\nClassification',
        'Class\nSize', 'Remarks',
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    entries = ScheduleEntry.objects.filter(
        tenant=tenant, academic_period=period,
    ).select_related(
        'course', 'course__department', 'faculty', 'room',
    ).prefetch_related('sections', 'sections__program').order_by('group_id', 'day_of_week')

    groups = defaultdict(list)
    for entry in entries:
        groups[str(entry.group_id)].append(entry)

    for group_id, group_entries in groups.items():
        first = group_entries[0]
        sections = first.sections.all()
        section_str = ', '.join(str(s) for s in sections)
        program_code = sections[0].program.code if sections else ''
        dept_code = first.course.department.code if first.course.department else ''

        days = [e.day_of_week for e in group_entries]
        time_ins = [_format_time(e.time_start) for e in group_entries]
        time_outs = [_format_time(e.time_end) for e in group_entries]
        load_classes = [e.load_classification for e in group_entries]

        if len(set(load_classes)) == 1:
            load_str = load_classes[0].replace('_', '-').title()
        else:
            load_parts = []
            for e in group_entries:
                day_name = e.day_of_week.capitalize()
                lc = e.load_classification.replace('_', '-').title()
                load_parts.append(f'{day_name} - {lc}')
            load_str = '\n'.join(load_parts)

        row = [
            program_code, dept_code, '',
            first.course.code, first.course.title,
            float(first.course.lec_units), float(first.course.lab_units),
            float(first.course.total_units), float(first.course.contact_hours),
            '', '',
            first.faculty.name if first.faculty else 'TBA',
            float(first.faculty_credits),
            '\n'.join(days), '\n'.join(time_ins), '\n'.join(time_outs),
            first.room.name, section_str, load_str,
            first.class_size, first.remarks,
        ]
        ws.append(row)

    return wb


def export_faculty_loading(tenant, period):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Faculty Loading'

    headers = [
        'Faculty Name', 'Employment Type', 'Total Units',
        'Regular Units', 'Overload Units', 'Built-in Units',
        'Part-time Units', 'Course Count', 'Section Count',
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    entries = ScheduleEntry.objects.filter(
        tenant=tenant, academic_period=period,
    ).exclude(faculty=None).select_related('course', 'faculty').prefetch_related('sections')

    faculty_data = defaultdict(lambda: {
        'employment_type': '', 'units': defaultdict(Decimal),
        'total': Decimal('0'), 'courses': set(), 'sections': set(),
    })

    for entry in entries:
        fac = entry.faculty
        data = faculty_data[fac.name]
        data['employment_type'] = fac.get_employment_type_display()
        units = entry.course.lec_units + entry.course.lab_units
        data['units'][entry.load_classification] += units
        data['total'] += units
        data['courses'].add(entry.course_id)
        for sec in entry.sections.all():
            data['sections'].add(sec.pk)

    for name in sorted(faculty_data.keys()):
        data = faculty_data[name]
        ws.append([
            name, data['employment_type'], float(data['total']),
            float(data['units'].get('REGULAR', 0)),
            float(data['units'].get('OVERLOAD', 0)),
            float(data['units'].get('BUILT_IN', 0)),
            float(data['units'].get('PART_TIME', 0)),
            len(data['courses']), len(data['sections']),
        ])

    return wb


def export_conflicts(tenant, period):
    """Full conflict report for presentation: one row per clash, with both the
    class and the class it conflicts with. A second sheet lists overloaded
    faculty (the accepted exemption)."""
    from .conflicts import detect_conflicts

    TYPE_LABELS = {
        'faculty': 'Faculty double-booked',
        'room': 'Room double-booked',
        'section': 'Section overlap',
    }

    entries = ScheduleEntry.objects.filter(
        tenant=tenant, academic_period=period,
    ).select_related('course', 'faculty', 'room').prefetch_related(
        'sections', 'sections__program',
    ).order_by('day_of_week', 'time_start')
    by_id = {e.pk: e for e in entries}

    def program_of(e):
        secs = list(e.sections.all())
        return secs[0].program.code if secs else ''

    def sections_of(e):
        return ', '.join(str(s) for s in e.sections.all())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Conflicts'

    headers = [
        'Class Code', 'Class Title', 'Program', 'Section(s)', 'Day',
        'Time Start', 'Time End', 'Room', 'Faculty', 'Load', 'Conflict Type',
        'Clashes With — Code', 'Clashes With — Title', 'Clashes With — Section(s)',
        'Clashes With — Day', 'Clashes With — Time', 'Clashes With — Room',
        'Clashes With — Faculty', 'Details',
    ]
    ws.append(headers)
    header_fill = PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid')
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill

    rows = 0
    seen_pairs = set()
    for e in entries:
        result = detect_conflicts(e)
        for h in result['hard']:
            # Each clash appears from both sides — list it once.
            pair = (h['type'], *sorted((e.pk, h.get('conflicting_entry_id') or 0)))
            if pair in seen_pairs:
                continue
            seen_pairs.add(pair)
            other = by_id.get(h.get('conflicting_entry_id'))
            ws.append([
                e.course.code, e.course.title, program_of(e), sections_of(e),
                e.day_of_week, _format_time(e.time_start), _format_time(e.time_end),
                e.room.name if e.room else '', e.faculty.name if e.faculty else 'TBA',
                e.load_classification.replace('_', '-').title(),
                TYPE_LABELS.get(h['type'], h['type']),
                other.course.code if other else '',
                other.course.title if other else '',
                sections_of(other) if other else '',
                other.day_of_week if other else '',
                f'{_format_time(other.time_start)}-{_format_time(other.time_end)}' if other else '',
                other.room.name if other and other.room else '',
                other.faculty.name if other and other.faculty else '',
                h.get('message', ''),
            ])
            rows += 1

    if rows == 0:
        ws.append(['No conflicts found.'])

    widths = [14, 30, 9, 18, 6, 11, 11, 14, 22, 10, 20, 14, 30, 18, 8, 22, 14, 22, 44]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    # --- Sheet 2: overloaded faculty (exemption) ---
    ws2 = wb.create_sheet('Faculty Overload')
    ws2.append(['Faculty', 'Assigned Units', 'Max Load', 'Over By'])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='E67E22', end_color='E67E22', fill_type='solid')

    totals = defaultdict(Decimal)
    fac_max = {}
    for e in entries:
        if e.faculty_id:
            totals[e.faculty.name] += (e.course.lec_units + e.course.lab_units)
            fac_max[e.faculty.name] = e.faculty.max_load_units
    over = [(n, u) for n, u in totals.items() if u > fac_max.get(n, Decimal('24'))]
    for name, units in sorted(over, key=lambda x: x[1], reverse=True):
        mx = fac_max.get(name, Decimal('24'))
        ws2.append([name, float(units), float(mx), float(units - mx)])
    if not over:
        ws2.append(['No overloaded faculty.'])
    for i, w in enumerate([30, 16, 12, 10], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = 'A2'

    return wb


def export_room_utilization(tenant, period):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Room Utilization'

    try:
        config = ScheduleConfig.objects.get(tenant=tenant, academic_period=period)
        operating_days = config.operating_days
        granularity = config.time_slot_granularity_minutes
        earliest = config.earliest_start_time
        latest = config.latest_end_time
    except ScheduleConfig.DoesNotExist:
        operating_days = ['MON', 'TUE', 'WED', 'THU', 'FRI']
        granularity = 60
        earliest = datetime.time(7, 0)
        latest = datetime.time(21, 0)

    slots = []
    current = datetime.datetime(2000, 1, 1, earliest.hour, earliest.minute)
    end = datetime.datetime(2000, 1, 1, latest.hour, latest.minute)
    while current < end:
        slots.append(current.time())
        current += datetime.timedelta(minutes=granularity)

    header = ['Room']
    for day in operating_days:
        for slot in slots:
            header.append(f'{day} {slot.strftime("%H:%M")}')
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    rooms = Room.objects.filter(tenant=tenant).order_by('building', 'floor', 'sequence_number', 'name')
    entries = ScheduleEntry.objects.filter(
        tenant=tenant, academic_period=period,
    ).select_related('course', 'room')

    occupied = {}
    for entry in entries:
        slot_start = datetime.datetime(2000, 1, 1, entry.time_start.hour, entry.time_start.minute)
        slot_end = datetime.datetime(2000, 1, 1, entry.time_end.hour, entry.time_end.minute)
        current = slot_start
        while current < slot_end:
            key = (entry.room_id, entry.day_of_week, current.time())
            occupied[key] = entry.course.code
            current += datetime.timedelta(minutes=granularity)

    fill_occupied = PatternFill(start_color='D4E6F1', end_color='D4E6F1', fill_type='solid')

    for room in rooms:
        row = [room.name]
        for day in operating_days:
            for slot in slots:
                code = occupied.get((room.pk, day, slot), '')
                row.append(code)
        ws.append(row)

        row_num = ws.max_row
        for col_idx in range(2, len(header) + 1):
            if ws.cell(row=row_num, column=col_idx).value:
                ws.cell(row=row_num, column=col_idx).fill = fill_occupied

    return wb
