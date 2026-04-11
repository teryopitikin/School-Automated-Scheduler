import datetime

import openpyxl
import pytest

from apps.scheduling.importers import (
    normalize_days,
    parse_load_classification,
    parse_time,
)

pytestmark = pytest.mark.django_db


# ---------------------------------------------------------------------------
# Shared fixtures
# ---------------------------------------------------------------------------
@pytest.fixture
def tenant():
    from apps.core.models import Tenant
    return Tenant.objects.create(name='Uni', slug='uni', status='ACTIVE')


@pytest.fixture
def period(tenant):
    from apps.scheduling.models import AcademicPeriod
    return AcademicPeriod.objects.create(
        tenant=tenant, name='1S 25-26', year_start=2025, year_end=2026,
        semester='1ST', status='DRAFT',
    )


# ---------------------------------------------------------------------------
# Unit tests — pure parser helpers
# ---------------------------------------------------------------------------
class TestNormalizeDays:
    def test_multiline(self):
        assert normalize_days('Tue\nThu') == ['TUE', 'THU']

    def test_abbreviation_mw(self):
        assert normalize_days('MW') == ['MON', 'WED']

    def test_abbreviation_tth(self):
        assert normalize_days('TTh') == ['TUE', 'THU']

    def test_abbreviation_mwf(self):
        assert normalize_days('MWF') == ['MON', 'WED', 'FRI']

    def test_single_day(self):
        assert normalize_days('Fri') == ['FRI']

    def test_full_single(self):
        assert normalize_days('Mon') == ['MON']

    def test_none(self):
        assert normalize_days(None) == []


class TestParseTime:
    def test_datetime_time(self):
        assert parse_time(datetime.time(9, 0)) == datetime.time(9, 0)

    def test_am_string(self):
        assert parse_time('9:00 AM') == datetime.time(9, 0)

    def test_pm_string(self):
        assert parse_time('1:00 PM') == datetime.time(13, 0)

    def test_short_am(self):
        assert parse_time('9:00A') == datetime.time(9, 0)

    def test_short_pm(self):
        assert parse_time('1:00P') == datetime.time(13, 0)

    def test_12pm(self):
        assert parse_time('12:00 PM') == datetime.time(12, 0)

    def test_12am(self):
        assert parse_time('12:00 AM') == datetime.time(0, 0)

    def test_range_extracts_start(self):
        assert parse_time('9:00A - 12:00P') == datetime.time(9, 0)

    def test_none(self):
        assert parse_time(None) is None


class TestParseLoadClassification:
    def test_single_value(self):
        result = parse_load_classification('Built-in', ['MON', 'WED'])
        assert result == {'MON': 'BUILT_IN', 'WED': 'BUILT_IN'}

    def test_per_day(self):
        result = parse_load_classification('Mon - Overload\nWed - Regular', ['MON', 'WED'])
        assert result == {'MON': 'OVERLOAD', 'WED': 'REGULAR'}

    def test_part_time(self):
        result = parse_load_classification('Part-time', ['MON'])
        assert result == {'MON': 'PART_TIME'}

    def test_overload(self):
        result = parse_load_classification('Overload', ['TUE', 'THU'])
        assert result == {'TUE': 'OVERLOAD', 'THU': 'OVERLOAD'}

    def test_none(self):
        result = parse_load_classification(None, ['MON'])
        assert result == {'MON': 'REGULAR'}


# ---------------------------------------------------------------------------
# Integration tests — import_excel
# ---------------------------------------------------------------------------
class TestImportExcel:
    def _make_workbook(self, rows):
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = [
            'Course', 'In-Charge', '', 'Course\nCode', 'Course Title',
            'Lec\nUnits', 'Lab\nUnits', 'Course\nUnits', 'Contact\nHours',
            '', '', 'Faculty', 'Faculty\nCredits', 'Day(s)', 'Time In',
            'Time Out', 'Room', 'Section', 'Load\nClassification',
            'Class\nSize', 'Remarks',
        ]
        ws.append(headers)
        for row in rows:
            ws.append(row)
        return wb

    def test_basic_import(self, tenant, period):
        from apps.scheduling.importers import import_excel
        from apps.scheduling.models import (
            Course, Faculty, Program, ScheduleEntry,
        )

        wb = self._make_workbook([
            ['BSA', 'Agri', '', 'CrSc 1', 'Crop Science', 3, 0, 3, 3,
             '', '', 'Ibao, Kristine', 3, 'MW', '7:30 AM', '9:00 AM',
             'Room 1', 'BSA 1-1', 'Regular', 25, ''],
        ])
        result = import_excel(wb, tenant, period)
        assert result['created'] == 2  # MW = 2 entries
        assert result['skipped'] == 0
        assert ScheduleEntry.objects.count() == 2
        assert Program.objects.filter(tenant=tenant, code='BSA').exists()
        assert Course.objects.filter(tenant=tenant, code='CrSc 1').exists()
        assert Faculty.objects.filter(tenant=tenant, name='Ibao, Kristine').exists()

    def test_multiline_days_times(self, tenant, period):
        from apps.scheduling.importers import import_excel

        wb = self._make_workbook([
            ['BSA', 'Agri', '', 'CrSc 1', 'Crop Science', 2, 1, 3, 5,
             '', '', 'Ibao, Kristine', 5, 'Tue\nThu',
             '9:00 AM\n10:00 AM', '12:00 PM\n12:00 PM',
             'AVR 1', 'BSA 1-1', 'Built-in', 33, ''],
        ])
        result = import_excel(wb, tenant, period)
        assert result['created'] == 2

    def test_tba_faculty(self, tenant, period):
        from apps.scheduling.importers import import_excel
        from apps.scheduling.models import ScheduleEntry

        wb = self._make_workbook([
            ['BSA', 'PE', '', 'PE 1', 'Physical Ed', 2, 0, 2, 2,
             '', '', 'TBA (PE)', 2, 'Fri', '9:00 AM', '11:00 AM',
             'Room 3', 'BSA 1-1', 'Overload', 27, ''],
        ])
        result = import_excel(wb, tenant, period)
        assert result['created'] == 1
        entry = ScheduleEntry.objects.first()
        assert entry.faculty is None
        assert entry.load_classification == 'OVERLOAD'

    def test_multi_section(self, tenant, period):
        from apps.scheduling.importers import import_excel
        from apps.scheduling.models import ScheduleEntry

        wb = self._make_workbook([
            ['BSA', 'PE', '', 'PE 1', 'Physical Ed', 2, 0, 2, 2,
             '', '', 'TBA', 2, 'Fri', '9:00 AM', '11:00 AM',
             'Room 3', 'BSA 1-1, BSF 1-1', 'Overload', 27, ''],
        ])
        result = import_excel(wb, tenant, period)
        entry = ScheduleEntry.objects.first()
        assert entry.sections.count() == 2

    def test_per_day_load_classification(self, tenant, period):
        from apps.scheduling.importers import import_excel
        from apps.scheduling.models import ScheduleEntry

        wb = self._make_workbook([
            ['BSA', 'Agri', '', 'CrSc 1', 'Crop Science', 2, 1, 3, 5,
             '', '', 'Dr. Smith', 5, 'Mon\nWed',
             '10:00 AM\n9:00 AM', '12:00 PM\n12:00 PM',
             'Room 1', 'BSA 1-1',
             'Mon - Overload\nWed - Regular', 25, ''],
        ])
        result = import_excel(wb, tenant, period)
        assert result['created'] == 2
        mon_entry = ScheduleEntry.objects.get(day_of_week='MON')
        wed_entry = ScheduleEntry.objects.get(day_of_week='WED')
        assert mon_entry.load_classification == 'OVERLOAD'
        assert wed_entry.load_classification == 'REGULAR'

    def test_vw_columns_flagged(self, tenant, period):
        from apps.scheduling.importers import import_excel

        wb = self._make_workbook([
            ['BSA', 'Agri', '', 'CrSc 1', 'Crop Science', 2, 1, 3, 5,
             '', '', 'Dr. Smith', 5, 'MW', '10:00 AM', '12:00 PM',
             'Room 1', 'BSA 1-1', 'Regular', 25, ''],
        ])
        ws = wb.active
        ws['V2'] = 'Tue\nThu'
        ws['W2'] = '9:00A - 12:00P\n10:00A - 12:00P'
        result = import_excel(wb, tenant, period)
        assert any('V/W' in w['reason'] for w in result['warnings'])
