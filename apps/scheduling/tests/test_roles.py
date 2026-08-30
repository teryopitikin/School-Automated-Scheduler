"""Role-based access: admins/registrars edit everything, department heads
(role DEPT_HEAD + managed_program_codes) edit only entries whose sections
belong to their programs and see only their programs' conflicts, viewers
are read-only. Import and wipe endpoints are admin-only."""
import datetime
import uuid

import pytest
from rest_framework.test import APIClient

from apps.core.models import Tenant, User
from apps.scheduling.models import (
    AcademicPeriod, Program, Department, Course, Section, Faculty, Room,
    ScheduleEntry,
)

pytestmark = pytest.mark.django_db


@pytest.fixture
def tenant():
    return Tenant.objects.create(name='Uni', slug='uni', status='ACTIVE')


@pytest.fixture
def period(tenant):
    return AcademicPeriod.objects.create(
        tenant=tenant, name='1S', year_start=2026, year_end=2027,
        semester='1ST', status='ACTIVE',
    )


@pytest.fixture
def dept(tenant):
    return Department.objects.create(tenant=tenant, code='GEN', name='General')


@pytest.fixture
def beed(tenant):
    return Program.objects.create(tenant=tenant, code='BEED', name='BEED')


@pytest.fixture
def bscrim(tenant):
    return Program.objects.create(tenant=tenant, code='BSCRIM', name='BSCRIM')


@pytest.fixture
def beed_sec(tenant, period, beed):
    return Section.objects.create(
        tenant=tenant, program=beed, academic_period=period,
        year_level=1, section_number=1,
    )


@pytest.fixture
def bscrim_sec(tenant, period, bscrim):
    return Section.objects.create(
        tenant=tenant, program=bscrim, academic_period=period,
        year_level=1, section_number=1,
    )


@pytest.fixture
def course(tenant, dept):
    return Course.objects.create(tenant=tenant, department=dept, code='GE 1', title='A')


@pytest.fixture
def faculty(tenant):
    return Faculty.objects.create(
        tenant=tenant, name='CRUZ, A', employment_type='FULL_TIME',
        max_load_units=999,
    )


@pytest.fixture
def room(tenant):
    return Room.objects.create(tenant=tenant, name='R1', room_type='LECTURE', capacity=40)


def make_user(tenant, username, role, codes=None, dept_codes=None, course_codes=None):
    return User.objects.create_user(
        username=username, password='pass', tenant=tenant, role=role,
        managed_program_codes=codes or [],
        managed_department_codes=dept_codes or [],
        managed_course_codes=course_codes or [],
    )


def client_for(user):
    c = APIClient()
    c.force_authenticate(user=user)
    return c


@pytest.fixture
def admin_client(tenant):
    return client_for(make_user(tenant, 'adm', 'ADMIN'))


@pytest.fixture
def registrar_client(tenant):
    return client_for(make_user(tenant, 'reg', 'REGISTRAR'))


@pytest.fixture
def head_client(tenant):
    return client_for(make_user(tenant, 'head', 'DEPT_HEAD', ['BEED']))


@pytest.fixture
def viewer_client(tenant):
    return client_for(make_user(tenant, 'view', 'VIEWER'))


def make_entry(tenant, period, course, faculty, room, sections, day='MON',
               start=(8, 0), end=(10, 0)):
    e = ScheduleEntry.objects.create(
        tenant=tenant, academic_period=period, course=course, faculty=faculty,
        room=room, day_of_week=day, time_start=datetime.time(*start),
        time_end=datetime.time(*end), group_id=uuid.uuid4(),
    )
    e.sections.set(sections)
    return e


@pytest.fixture
def beed_entry(tenant, period, course, faculty, room, beed_sec):
    return make_entry(tenant, period, course, faculty, room, [beed_sec])


@pytest.fixture
def bscrim_entry(tenant, period, course, faculty, room, bscrim_sec):
    return make_entry(tenant, period, course, faculty, room, [bscrim_sec],
                      day='TUE')


class TestEntryWritePermissions:
    def test_viewer_cannot_create(self, viewer_client, period, course, room, beed_sec):
        resp = viewer_client.post('/api/scheduler/schedules/', {
            'academic_period': period.pk, 'course': course.pk, 'room': room.pk,
            'time_start': '13:00', 'time_end': '14:00', 'days': ['WED'],
            'sections': [beed_sec.pk],
        }, format='json')
        assert resp.status_code == 403

    def test_viewer_cannot_edit_or_delete(self, viewer_client, beed_entry):
        assert viewer_client.patch(
            f'/api/scheduler/schedules/{beed_entry.pk}/',
            {'day_of_week': 'WED'}, format='json').status_code == 403
        assert viewer_client.delete(
            f'/api/scheduler/schedules/{beed_entry.pk}/').status_code == 403

    def test_viewer_can_read(self, viewer_client, beed_entry):
        assert viewer_client.get('/api/scheduler/schedules/').status_code == 200

    def test_registrar_edits_anything(self, registrar_client, beed_entry, bscrim_entry):
        for e in (beed_entry, bscrim_entry):
            resp = registrar_client.patch(
                f'/api/scheduler/schedules/{e.pk}/',
                {'time_start': '15:00', 'time_end': '16:00'}, format='json')
            assert resp.status_code == 200

    def test_admin_edits_anything(self, admin_client, bscrim_entry):
        resp = admin_client.patch(
            f'/api/scheduler/schedules/{bscrim_entry.pk}/',
            {'time_start': '15:00', 'time_end': '16:00'}, format='json')
        assert resp.status_code == 200

    def test_head_edits_own_program_entry(self, head_client, beed_entry):
        resp = head_client.patch(
            f'/api/scheduler/schedules/{beed_entry.pk}/',
            {'time_start': '15:00', 'time_end': '16:00'}, format='json')
        assert resp.status_code == 200

    def test_head_cannot_edit_foreign_entry(self, head_client, bscrim_entry):
        resp = head_client.patch(
            f'/api/scheduler/schedules/{bscrim_entry.pk}/',
            {'time_start': '15:00', 'time_end': '16:00'}, format='json')
        assert resp.status_code == 403

    def test_head_deletes_own_not_foreign(self, head_client, beed_entry, bscrim_entry):
        assert head_client.delete(
            f'/api/scheduler/schedules/{bscrim_entry.pk}/').status_code == 403
        assert head_client.delete(
            f'/api/scheduler/schedules/{beed_entry.pk}/').status_code == 204

    def test_head_edits_cotaught_entry(self, head_client, tenant, period, course,
                                       faculty, room, beed_sec, bscrim_sec):
        e = make_entry(tenant, period, course, faculty, room,
                       [beed_sec, bscrim_sec], day='WED')
        resp = head_client.patch(
            f'/api/scheduler/schedules/{e.pk}/',
            {'time_start': '15:00', 'time_end': '16:00'}, format='json')
        assert resp.status_code == 200

    def test_head_creates_for_own_section(self, head_client, period, course, room, beed_sec):
        resp = head_client.post('/api/scheduler/schedules/', {
            'academic_period': period.pk, 'course': course.pk, 'room': room.pk,
            'time_start': '13:00', 'time_end': '14:00', 'days': ['WED'],
            'sections': [beed_sec.pk],
        }, format='json')
        assert resp.status_code == 201

    def test_head_cannot_create_for_foreign_section(self, head_client, period,
                                                    course, room, bscrim_sec):
        resp = head_client.post('/api/scheduler/schedules/', {
            'academic_period': period.pk, 'course': course.pk, 'room': room.pk,
            'time_start': '13:00', 'time_end': '14:00', 'days': ['WED'],
            'sections': [bscrim_sec.pk],
        }, format='json')
        assert resp.status_code == 403

    def test_head_cannot_create_sectionless(self, head_client, period, course, room):
        resp = head_client.post('/api/scheduler/schedules/', {
            'academic_period': period.pk, 'course': course.pk, 'room': room.pk,
            'time_start': '13:00', 'time_end': '14:00', 'days': ['WED'],
            'sections': [],
        }, format='json')
        assert resp.status_code == 403

    def test_head_group_actions_scoped(self, head_client, beed_entry, bscrim_entry):
        assert head_client.post(
            f'/api/scheduler/schedules/{bscrim_entry.pk}/edit-group/',
            {'time_start': '15:00', 'time_end': '16:00'}, format='json').status_code == 403
        assert head_client.post(
            f'/api/scheduler/schedules/{bscrim_entry.pk}/delete-group/').status_code == 403
        assert head_client.post(
            f'/api/scheduler/schedules/{beed_entry.pk}/edit-group/',
            {'time_start': '15:00', 'time_end': '16:00'}, format='json').status_code == 200


class TestConflictScoping:
    @pytest.fixture
    def clashes(self, tenant, period, course, room, beed_sec, bscrim_sec):
        """One faculty clash inside BEED, one inside BSCRIM (separate teachers)."""
        f1 = Faculty.objects.create(tenant=tenant, name='F1', max_load_units=999)
        f2 = Faculty.objects.create(tenant=tenant, name='F2', max_load_units=999)
        b1 = make_entry(tenant, period, course, f1, room, [beed_sec], day='MON')
        b2 = make_entry(tenant, period, course, f1, room, [beed_sec], day='MON',
                        start=(9, 0), end=(11, 0))
        c1 = make_entry(tenant, period, course, f2, room, [bscrim_sec], day='TUE')
        c2 = make_entry(tenant, period, course, f2, room, [bscrim_sec], day='TUE',
                        start=(9, 0), end=(11, 0))
        return b1, b2, c1, c2

    def test_admin_sees_all_conflicts(self, admin_client, period, clashes):
        resp = admin_client.get(
            f'/api/scheduler/schedules/conflicts/?academic_period={period.pk}')
        assert {i['entry_id'] for i in resp.data} == {e.pk for e in clashes}

    def test_head_sees_only_own_program_conflicts(self, head_client, period, clashes):
        b1, b2, c1, c2 = clashes
        resp = head_client.get(
            f'/api/scheduler/schedules/conflicts/?academic_period={period.pk}')
        assert {i['entry_id'] for i in resp.data} == {b1.pk, b2.pk}

    def test_cross_program_conflict_visible_to_both_heads(
            self, tenant, period, course, room, beed_sec, bscrim_sec):
        f = Faculty.objects.create(tenant=tenant, name='F3', max_load_units=999)
        e1 = make_entry(tenant, period, course, f, room, [beed_sec], day='FRI')
        e2 = make_entry(tenant, period, course, f, room, [bscrim_sec], day='FRI',
                        start=(9, 0), end=(11, 0))
        for codes, expected in ((['BEED'], {e1.pk, e2.pk}), (['BSCRIM'], {e1.pk, e2.pk})):
            client = client_for(make_user(tenant, f'h{codes[0]}', 'DEPT_HEAD', codes))
            resp = client.get(
                f'/api/scheduler/schedules/conflicts/?academic_period={period.pk}')
            got = {i['entry_id'] for i in resp.data}
            assert got == expected, codes


class TestImportWipePermissions:
    def test_wipe_admin_only(self, admin_client, registrar_client, head_client,
                             viewer_client, beed_entry):
        for client in (registrar_client, head_client, viewer_client):
            assert client.post('/api/scheduler/wipe-schedule/').status_code == 403
        resp = admin_client.post('/api/scheduler/wipe-schedule/')
        assert resp.status_code == 200
        assert resp.data['wiped']['ScheduleEntry'] == 1
        assert ScheduleEntry.objects.count() == 0
        assert Program.objects.count() == 0

    def test_wipe_preserves_users_and_periods(self, admin_client, period, beed_entry):
        admin_client.post('/api/scheduler/wipe-schedule/')
        assert AcademicPeriod.objects.filter(pk=period.pk).exists()
        assert User.objects.filter(username='adm').exists()

    def test_full_export_import_admin_only(self, registrar_client, head_client, viewer_client):
        for client in (registrar_client, head_client, viewer_client):
            resp = client.post('/api/scheduler/import-full-export/', {})
            assert resp.status_code == 403

    def test_cleaned_import_admin_only(self, registrar_client):
        resp = registrar_client.post('/api/scheduler/import/', {})
        assert resp.status_code == 403


class TestDepartmentAndCourseScoping:
    """Heads can also be assigned departments (scope = entries whose course
    belongs to the department) and individual courses."""

    @pytest.fixture
    def crim_dept(self, tenant):
        return Department.objects.create(tenant=tenant, code='CRIM', name='Criminology')

    @pytest.fixture
    def crim_course(self, tenant, crim_dept):
        return Course.objects.create(tenant=tenant, department=crim_dept,
                                     code='CLJ 1', title='CLJ')

    @pytest.fixture
    def crim_entry(self, tenant, period, crim_course, faculty, room, bscrim_sec):
        return make_entry(tenant, period, crim_course, faculty, room,
                          [bscrim_sec], day='THU')

    def test_department_assignee_edits_matching_course_entry(
            self, tenant, crim_entry):
        client = client_for(make_user(tenant, 'dhead', 'DEPT_HEAD',
                                      dept_codes=['CRIM']))
        resp = client.patch(f'/api/scheduler/schedules/{crim_entry.pk}/',
                            {'time_start': '15:00', 'time_end': '16:00'},
                            format='json')
        assert resp.status_code == 200

    def test_department_assignee_cannot_edit_other_department(
            self, tenant, crim_entry, beed_entry):
        client = client_for(make_user(tenant, 'dhead2', 'DEPT_HEAD',
                                      dept_codes=['CRIM']))
        resp = client.patch(f'/api/scheduler/schedules/{beed_entry.pk}/',
                            {'time_start': '15:00', 'time_end': '16:00'},
                            format='json')
        assert resp.status_code == 403

    def test_course_assignee_edits_that_course_only(
            self, tenant, crim_entry, beed_entry):
        client = client_for(make_user(tenant, 'chead', 'DEPT_HEAD',
                                      course_codes=['CLJ 1']))
        ok = client.patch(f'/api/scheduler/schedules/{crim_entry.pk}/',
                          {'time_start': '15:00', 'time_end': '16:00'},
                          format='json')
        no = client.patch(f'/api/scheduler/schedules/{beed_entry.pk}/',
                          {'time_start': '17:00', 'time_end': '18:00'},
                          format='json')
        assert (ok.status_code, no.status_code) == (200, 403)

    def test_course_assignee_can_create_for_that_course(
            self, tenant, period, crim_course, room, bscrim_sec):
        client = client_for(make_user(tenant, 'chead2', 'DEPT_HEAD',
                                      course_codes=['CLJ 1']))
        resp = client.post('/api/scheduler/schedules/', {
            'academic_period': period.pk, 'course': crim_course.pk,
            'room': room.pk, 'time_start': '13:00', 'time_end': '14:00',
            'days': ['WED'], 'sections': [bscrim_sec.pk],
        }, format='json')
        assert resp.status_code == 201

    def test_department_scoped_conflicts(self, tenant, period, crim_course,
                                         room, bscrim_sec, course, beed_sec):
        f = Faculty.objects.create(tenant=tenant, name='FX', max_load_units=999)
        c1 = make_entry(tenant, period, crim_course, f, room, [bscrim_sec], day='THU')
        c2 = make_entry(tenant, period, crim_course, f, room, [bscrim_sec], day='THU',
                        start=(9, 0), end=(11, 0))
        f2 = Faculty.objects.create(tenant=tenant, name='FY', max_load_units=999)
        make_entry(tenant, period, course, f2, room, [beed_sec], day='MON')
        make_entry(tenant, period, course, f2, room, [beed_sec], day='MON',
                   start=(9, 0), end=(11, 0))
        client = client_for(make_user(tenant, 'dhead3', 'DEPT_HEAD',
                                      dept_codes=['CRIM']))
        resp = client.get(
            f'/api/scheduler/schedules/conflicts/?academic_period={period.pk}')
        assert {i['entry_id'] for i in resp.data} == {c1.pk, c2.pk}


class TestMetadataImport:
    """Admin-only upsert import of Departments / Programs / Courses from a
    workbook with any subset of those sheets. Never touches the schedule."""

    def _workbook(self):
        import io

        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Departments'
        ws.append(['Code', 'Name'])
        ws.append(['CRIM', 'Criminology'])
        ws2 = wb.create_sheet('Programs')
        ws2.append(['Code', 'Name'])
        ws2.append(['BSCRIM', 'BS Criminology'])
        ws3 = wb.create_sheet('Courses')
        ws3.append(['Code', 'Title', 'Department', 'Lec', 'Lab', 'Has Lab'])
        ws3.append(['CLJ 1', 'CLJ', 'CRIM', 3, 0, 'No'])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = 'meta.xlsx'
        return buf

    def test_admin_only(self, registrar_client, head_client, viewer_client):
        for client in (registrar_client, head_client, viewer_client):
            assert client.post('/api/scheduler/import-metadata/', {}).status_code == 403

    def test_creates_and_updates_without_touching_schedule(
            self, admin_client, tenant, beed_entry, bscrim):
        resp = admin_client.post('/api/scheduler/import-metadata/',
                                 {'file': self._workbook()}, format='multipart')
        assert resp.status_code == 201, resp.data
        assert Department.objects.filter(tenant=tenant, code='CRIM').exists()
        assert Course.objects.filter(tenant=tenant, code='CLJ 1').exists()
        # existing program updated, not duplicated
        assert Program.objects.filter(tenant=tenant, code='BSCRIM').count() == 1
        assert Program.objects.get(tenant=tenant, code='BSCRIM').name == 'BS Criminology'
        # schedule untouched
        assert ScheduleEntry.objects.filter(pk=beed_entry.pk).exists()
        assert resp.data['departments'] == {'created': 1, 'updated': 0}
        assert resp.data['programs'] == {'created': 0, 'updated': 1}
        assert resp.data['courses'] == {'created': 1, 'updated': 0}


class TestAssistantExecutePermissions:
    def test_viewer_cannot_execute(self, viewer_client, beed_entry):
        resp = viewer_client.post('/api/scheduler/assistant/execute/', {
            'action': {'type': 'move_class',
                       'payload': {'entry_id': beed_entry.pk, 'day_of_week': 'WED'}},
        }, format='json')
        assert resp.status_code == 403

    def test_head_cannot_move_foreign_entry(self, head_client, bscrim_entry):
        resp = head_client.post('/api/scheduler/assistant/execute/', {
            'action': {'type': 'move_class',
                       'payload': {'entry_id': bscrim_entry.pk, 'day_of_week': 'WED'}},
        }, format='json')
        assert resp.status_code == 403

    def test_head_moves_own_entry(self, head_client, beed_entry):
        resp = head_client.post('/api/scheduler/assistant/execute/', {
            'action': {'type': 'move_class',
                       'payload': {'entry_id': beed_entry.pk, 'day_of_week': 'WED'}},
        }, format='json')
        assert resp.status_code == 200

    def test_head_cannot_delete_foreign_entry(self, head_client, bscrim_entry):
        resp = head_client.post('/api/scheduler/assistant/execute/', {
            'action': {'type': 'delete_class',
                       'payload': {'entry_id': bscrim_entry.pk}},
        }, format='json')
        assert resp.status_code == 403

    def test_head_cannot_add_for_foreign_section(self, head_client, period, course,
                                                 room, bscrim_sec):
        resp = head_client.post('/api/scheduler/assistant/execute/', {
            'action': {'type': 'add_class', 'payload': {
                'academic_period': period.pk, 'course': course.pk, 'room': room.pk,
                'time_start': '13:00', 'time_end': '14:00', 'days': ['WED'],
                'sections': [bscrim_sec.pk],
            }},
        }, format='json')
        assert resp.status_code == 403


class TestMetadataWritePermissions:
    """Data pages: Admin/Registrar edit everything; a DEPT_HEAD may only
    edit Courses within their assigned departments (or assigned courses)
    and Sections of their assigned programs; everything else — and every
    write for viewers — is read-only."""

    @pytest.fixture
    def crim_dept(self, tenant):
        return Department.objects.create(tenant=tenant, code='CRIM', name='Criminology')

    @pytest.fixture
    def crim_head_client(self, tenant):
        return client_for(make_user(tenant, 'crimhead', 'DEPT_HEAD',
                                    codes=['BSCRIM'], dept_codes=['CRIM']))

    def test_viewer_cannot_write_metadata(self, viewer_client, course, room):
        assert viewer_client.post('/api/scheduler/courses/', {}).status_code == 403
        assert viewer_client.patch(f'/api/scheduler/courses/{course.pk}/',
                                   {'title': 'X'}, format='json').status_code == 403
        assert viewer_client.delete(f'/api/scheduler/rooms/{room.pk}/').status_code == 403

    def test_head_edits_course_in_own_department(self, crim_head_client, tenant, crim_dept):
        c = Course.objects.create(tenant=tenant, department=crim_dept, code='CLJ 1', title='x')
        resp = crim_head_client.patch(f'/api/scheduler/courses/{c.pk}/',
                                      {'title': 'CLJ One'}, format='json')
        assert resp.status_code == 200

    def test_head_cannot_edit_foreign_course(self, crim_head_client, course):
        resp = crim_head_client.patch(f'/api/scheduler/courses/{course.pk}/',
                                      {'title': 'X'}, format='json')
        assert resp.status_code == 403

    def test_assigned_course_editable_even_outside_dept(self, tenant, course):
        client = client_for(make_user(tenant, 'chead', 'DEPT_HEAD',
                                      course_codes=['GE 1']))
        resp = client.patch(f'/api/scheduler/courses/{course.pk}/',
                            {'title': 'GE One'}, format='json')
        assert resp.status_code == 200

    def test_head_creates_course_only_in_own_department(self, crim_head_client,
                                                        crim_dept, dept):
        ok = crim_head_client.post('/api/scheduler/courses/', {
            'department': crim_dept.pk, 'code': 'CLJ 9', 'title': 'New',
            'lec_units': 3, 'lab_units': 0, 'contact_hours': 3, 'has_lab': False,
        }, format='json')
        no = crim_head_client.post('/api/scheduler/courses/', {
            'department': dept.pk, 'code': 'GEN 9', 'title': 'New',
            'lec_units': 3, 'lab_units': 0, 'contact_hours': 3, 'has_lab': False,
        }, format='json')
        assert (ok.status_code, no.status_code) == (201, 403)

    def test_head_creates_section_only_in_own_program(self, crim_head_client, period,
                                                      beed, bscrim):
        ok = crim_head_client.post('/api/scheduler/sections/', {
            'program': bscrim.pk, 'academic_period': period.pk,
            'year_level': 2, 'section_number': 1,
        }, format='json')
        no = crim_head_client.post('/api/scheduler/sections/', {
            'program': beed.pk, 'academic_period': period.pk,
            'year_level': 2, 'section_number': 1,
        }, format='json')
        assert (ok.status_code, no.status_code) == (201, 403)

    def test_head_edits_own_program_section_not_foreign(self, crim_head_client,
                                                        beed_sec, bscrim_sec):
        ok = crim_head_client.patch(f'/api/scheduler/sections/{bscrim_sec.pk}/',
                                    {'section_number': 5}, format='json')
        no = crim_head_client.patch(f'/api/scheduler/sections/{beed_sec.pk}/',
                                    {'section_number': 5}, format='json')
        assert (ok.status_code, no.status_code) == (200, 403)

    def test_head_read_only_on_other_metadata(self, crim_head_client, room, faculty,
                                              crim_dept, bscrim):
        assert crim_head_client.get('/api/scheduler/rooms/').status_code == 200
        assert crim_head_client.patch(f'/api/scheduler/rooms/{room.pk}/',
                                      {'capacity': 50}, format='json').status_code == 403
        assert crim_head_client.patch(f'/api/scheduler/faculty/{faculty.pk}/',
                                      {'name': 'X'}, format='json').status_code == 403
        # even their own department/program records are read-only
        assert crim_head_client.patch(f'/api/scheduler/departments/{crim_dept.pk}/',
                                      {'name': 'X'}, format='json').status_code == 403
        assert crim_head_client.patch(f'/api/scheduler/programs/{bscrim.pk}/',
                                      {'name': 'X'}, format='json').status_code == 403

    def test_registrar_still_edits_everything(self, registrar_client, course, room):
        assert registrar_client.patch(f'/api/scheduler/courses/{course.pk}/',
                                      {'title': 'Y'}, format='json').status_code == 200
        assert registrar_client.patch(f'/api/scheduler/rooms/{room.pk}/',
                                      {'capacity': 45}, format='json').status_code == 200
