import datetime
import uuid

import pytest
import openpyxl

from apps.core.models import Tenant
from apps.scheduling.models import (
    AcademicPeriod, Program, Department, Course, Section,
    Faculty, Room, ScheduleEntry, ScheduleConfig,
)
from apps.scheduling.exporters import export_schedule, export_faculty_loading, export_room_utilization

pytestmark = pytest.mark.django_db


@pytest.fixture
def tenant():
    return Tenant.objects.create(name='Uni', slug='uni', status='ACTIVE')


@pytest.fixture
def period(tenant):
    return AcademicPeriod.objects.create(
        tenant=tenant, name='1S 25-26', year_start=2025, year_end=2026,
        semester='1ST', status='DRAFT',
    )


@pytest.fixture
def config(tenant, period):
    return ScheduleConfig.objects.create(
        tenant=tenant, academic_period=period,
        earliest_start_time=datetime.time(7, 0),
        latest_end_time=datetime.time(12, 0),
        time_slot_granularity_minutes=60,
        operating_days=['MON', 'TUE', 'WED', 'THU', 'FRI'],
    )


@pytest.fixture
def full_schedule(tenant, period):
    dept = Department.objects.create(tenant=tenant, code='Agri', name='Agriculture')
    course = Course.objects.create(
        tenant=tenant, department=dept, code='CrSc 1', title='Crop Science 1',
        lec_units=3, lab_units=0, contact_hours=3, has_lab=False,
    )
    faculty = Faculty.objects.create(
        tenant=tenant, name='Ibao, Kristine', employment_type='FULL_TIME',
        priority_level=5, max_load_units=24,
    )
    room = Room.objects.create(
        tenant=tenant, name='Room 101', room_type='LECTURE', capacity=40,
        building='Main', floor=1, sequence_number=1,
    )
    prog = Program.objects.create(tenant=tenant, code='BSA', name='BSA')
    sec = Section.objects.create(
        tenant=tenant, program=prog, academic_period=period,
        year_level=1, section_number=1,
    )
    group = uuid.uuid4()
    for day in ['MON', 'WED']:
        entry = ScheduleEntry.objects.create(
            tenant=tenant, academic_period=period, course=course,
            faculty=faculty, room=room, day_of_week=day,
            time_start=datetime.time(8, 0), time_end=datetime.time(9, 30),
            group_id=group, entry_type='LECTURE',
            load_classification='REGULAR', class_size=25,
            faculty_credits=3,
        )
        entry.sections.set([sec])
    return {'course': course, 'faculty': faculty, 'room': room, 'section': sec}


class TestExportSchedule:
    def test_basic_export(self, tenant, period, full_schedule):
        wb = export_schedule(tenant, period)
        ws = wb.active
        assert ws['D1'].value == 'Course\nCode'
        assert ws['D2'].value == 'CrSc 1'

    def test_multiday_collapsed(self, tenant, period, full_schedule):
        wb = export_schedule(tenant, period)
        ws = wb.active
        data_rows = [row for row in ws.iter_rows(min_row=2, values_only=True) if row[3]]
        assert len(data_rows) == 1  # 2 entries collapsed into 1 row

    def test_empty_period(self, tenant, period):
        wb = export_schedule(tenant, period)
        ws = wb.active
        data_rows = [row for row in ws.iter_rows(min_row=2, values_only=True) if row[3]]
        assert len(data_rows) == 0


class TestExportFacultyLoading:
    def test_basic(self, tenant, period, full_schedule):
        wb = export_faculty_loading(tenant, period)
        ws = wb.active
        assert ws['A1'].value == 'Faculty Name'
        assert ws['A2'].value == 'Ibao, Kristine'

    def test_empty(self, tenant, period):
        wb = export_faculty_loading(tenant, period)
        ws = wb.active
        data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
        assert len(data_rows) == 0


class TestExportRoomUtilization:
    def test_basic(self, tenant, period, config, full_schedule):
        wb = export_room_utilization(tenant, period)
        ws = wb.active
        assert ws['A1'].value == 'Room'
        assert ws['A2'].value == 'Room 101'
