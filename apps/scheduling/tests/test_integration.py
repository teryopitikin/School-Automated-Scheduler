import datetime
from pathlib import Path

import openpyxl
import pytest

from apps.core.models import Tenant
from apps.scheduling.importers import import_excel
from apps.scheduling.exporters import export_schedule, export_faculty_loading
from apps.scheduling.stats import compute_stats
from apps.scheduling.models import (
    AcademicPeriod, ScheduleConfig, ScheduleEntry, Faculty, Course, Program,
)

pytestmark = pytest.mark.django_db

REFERENCE_EXCEL = Path('/home/classify/Desktop/Errors/NH Faculty Loading 1S 25-26.xlsx')


@pytest.fixture
def tenant():
    return Tenant.objects.create(name='NH', slug='nh', status='ACTIVE')


@pytest.fixture
def period(tenant):
    return AcademicPeriod.objects.create(
        tenant=tenant, name='1S 25-26', year_start=2025, year_end=2026,
        semester='1ST', status='DRAFT',
    )


@pytest.fixture
def config(tenant, period):
    return ScheduleConfig.objects.create(
        tenant=tenant, academic_period=period,
        earliest_start_time=datetime.time(7, 0),
        latest_end_time=datetime.time(21, 0),
        time_slot_granularity_minutes=30,
        operating_days=['MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT'],
    )


@pytest.mark.skipif(not REFERENCE_EXCEL.exists(), reason='Reference Excel not found')
class TestFullIntegration:
    def test_import_reference_excel(self, tenant, period, config):
        wb = openpyxl.load_workbook(REFERENCE_EXCEL)
        result = import_excel(wb, tenant, period)

        assert result['created'] > 0
        assert ScheduleEntry.objects.filter(tenant=tenant, academic_period=period).count() > 0
        assert Faculty.objects.filter(tenant=tenant).count() > 0
        assert Course.objects.filter(tenant=tenant).count() > 0
        assert Program.objects.filter(tenant=tenant).count() > 0

    def test_stats_after_import(self, tenant, period, config):
        wb = openpyxl.load_workbook(REFERENCE_EXCEL)
        import_excel(wb, tenant, period)

        stats = compute_stats(tenant, period)
        assert stats['summary']['total_courses'] > 0
        assert stats['summary']['faculty_count'] > 0
        assert len(stats['faculty_breakdown']) > 0
        assert len(stats['program_progress']) > 0

    def test_export_after_import(self, tenant, period, config):
        wb = openpyxl.load_workbook(REFERENCE_EXCEL)
        import_excel(wb, tenant, period)

        export_wb = export_schedule(tenant, period)
        ws = export_wb.active
        data_rows = [row for row in ws.iter_rows(min_row=2, values_only=True) if row[3]]
        assert len(data_rows) > 0

        fac_wb = export_faculty_loading(tenant, period)
        fac_ws = fac_wb.active
        fac_rows = [row for row in fac_ws.iter_rows(min_row=2, values_only=True) if row[0]]
        assert len(fac_rows) > 0
