"""Convert a per-program 'SCHEDULE OF CLASSES' .docx (one table per year
level: DAY | TIME | COURSE CODE | COURSE TITLE | UNIT/S | ROOM | INSTRUCTOR)
into the cleaned-import workbook format that
apps/scheduling/cleaned_importer.import_cleaned consumes.

Usage:
    python scripts/docx_to_cleaned.py "<schedule.docx>" <PROGRAM_CODE> [out.xlsx]

e.g. python scripts/docx_to_cleaned.py \
        "~/Downloads/BSED SST_S.Y._ 2026-2027_FIRST SEMESTER.docx" BSED-SST

Each year-level table becomes section <PROGRAM_CODE>-<n>A. Appends to the
output workbook if it already exists, so several program files can be
accumulated into one import file.
"""
import datetime
import os
import re
import sys

import docx
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

SEMESTER = 'First Semester 2026-2027'

DAY_LABEL = {'MON': 'M', 'TUE': 'T', 'WED': 'W', 'THU': 'TH',
             'FRI': 'F', 'SAT': 'Sat', 'SUN': 'Sun'}
DAY_ORDER = ['MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT', 'SUN']

YEAR_WORDS = {'FIRST': 1, 'SECOND': 2, 'THIRD': 3, 'FOURTH': 4, 'FIFTH': 5}

# Targeted source-data fixes
CODE_FIXES = {
    'PROD ED 107': 'PROF ED 107',   # recurring PROD/PROF typo in BSED sheets
}
NAME_FIXES = {
    'JUN CARLO FILEPE': 'JUN CARLO FELIPE',   # BEED 3rd-yr typo
    'MICUTUAN. E.': 'MICUTUAN, E.',           # stray period
}

FULLNAME = {
    'MONDAY': 'MON', 'MON': 'MON',
    'TUESDAY': 'TUE', 'TUES': 'TUE', 'TUE': 'TUE',
    'WEDNESDAY': 'WED', 'WED': 'WED',
    'THURSDAY': 'THU', 'THURS': 'THU', 'THUR': 'THU', 'THU': 'THU', 'TH': 'THU',
    'FRIDAY': 'FRI', 'FRI': 'FRI',
    'SATURDAY': 'SAT', 'SAT': 'SAT',
    'SUNDAY': 'SUN', 'SUN': 'SUN',
}


def _compact_days(s):
    """'MWF' -> ['MON','WED','FRI']; 'TTH' -> ['TUE','THU']; 'MTWTH' etc."""
    out, i = [], 0
    while i < len(s):
        if s[i:i + 3] in ('SAT', 'SUN'):
            out.append(s[i:i + 3]); i += 3; continue
        if s[i:i + 2] == 'TH':
            out.append('THU'); i += 2; continue
        one = {'M': 'MON', 'T': 'TUE', 'W': 'WED', 'F': 'FRI', 'S': 'SAT'}.get(s[i])
        if one is None:
            return None
        out.append(one); i += 1
    return out or None


def parse_day_string(s):
    """'MWF', 'M, T, F', 'Monday', 'MTTH', 'FRI', 'M-F' -> day-code list."""
    s = s.strip().upper().replace('.', '')
    # day range like 'M-F' / 'MON-FRI' -> expand across the week
    m = re.fullmatch(r'([A-Z]+)\s*-\s*([A-Z]+)', s)
    if m:
        a = FULLNAME.get(m.group(1)) or (_compact_days(m.group(1)) or [None])[0]
        b = FULLNAME.get(m.group(2)) or (_compact_days(m.group(2)) or [None])[0]
        if a in DAY_ORDER and b in DAY_ORDER and DAY_ORDER.index(a) < DAY_ORDER.index(b):
            return DAY_ORDER[DAY_ORDER.index(a):DAY_ORDER.index(b) + 1]
    out = []
    for tok in re.split(r'[,\s&/]+', s):
        if not tok:
            continue
        if tok in FULLNAME:
            out.append(FULLNAME[tok])
            continue
        sub = _compact_days(re.sub(r'[^A-Z]', '', tok))
        if sub is None:
            return None
        out.extend(sub)
    return out or None


def _school_hour(h):
    """Bare hour -> 24h in a school-day context (1-6 -> PM)."""
    if h == 12:
        return 12
    if 1 <= h <= 6:
        return h + 12
    return h


def _to24(h, m, mer):
    if mer == 'NN':
        return datetime.time(12, m)
    if mer == 'PM' and h != 12:
        h += 12
    elif mer == 'AM' and h == 12:
        h = 0
    return datetime.time(h % 24, m)


def parse_time_range(s):
    """'7:00-8:00AM', '10:00-11:00', '9:00-12:00AM' -> (start, end) times."""
    toks = re.findall(r'(\d{1,2})(?::(\d{2}))?\s*(AM|PM|NN)?', s.upper())
    toks = [t for t in toks if t[0]]
    if len(toks) < 2:
        return None
    (h1, m1, mer1), (h2, m2, mer2) = toks[0], toks[1]
    h1, h2 = int(h1), int(h2)
    m1 = int(m1) if m1 else 0
    m2 = int(m2) if m2 else 0
    mer1 = mer1 or None
    mer2 = mer2 or None
    if mer1 is None and mer2 is not None:
        mer1 = ('AM' if h1 < 12 else 'PM') if mer2 == 'NN' else mer2
    if mer2 is None and mer1 is not None:
        mer2 = mer1 if mer1 != 'NN' else 'PM'
    if mer1 is None and mer2 is None:
        start = datetime.time(_school_hour(h1), m1)
        end = datetime.time(_school_hour(h2), m2)
    else:
        start, end = _to24(h1, m1, mer1), _to24(h2, m2, mer2)
        if start >= end and mer1 == mer2:   # '9:00-12:00AM' style: fix the start
            alt = _to24(h1, m1, 'AM' if mer1 == 'PM' else 'PM')
            if alt < end:
                start = alt
    # classes never run at midnight; 00:00 is really noon
    if start.hour == 0:
        start = datetime.time(12, start.minute)
    if end.hour == 0:
        end = datetime.time(12, end.minute)
    if start >= end and end.hour < 12:      # end still earlier -> it's PM
        end = datetime.time(end.hour + 12, end.minute)
    return start, end


def fmt(t):
    return t.strftime('%I:%M %p').lstrip('0')


def clean_cell(text):
    return re.sub(r'\s+', ' ', text.replace('\n', ' ')).strip()


def norm_room(raw):
    raw = clean_cell(raw)
    if re.fullmatch(r'\d+', raw):
        raw = f'Room {raw}'
    raw = re.sub(r'(?i)^room\b', 'Room', raw)
    # 'Room 016' / 'Room 07' -> 'Room 16' / 'Room 7' so one room = one record
    return re.sub(r'^Room 0+(\d)', r'Room \1', raw)


def year_from_header(text):
    for word, n in YEAR_WORDS.items():
        if word in text.upper():
            return n
    return None


HEADERS = ['Subject Code', 'Descriptive Title', 'Units', 'Semester',
           'Day', 'Time', 'Time Start', 'Time End', 'Meeting',
           'Room Number', 'Faculty Name', 'Course', 'Year Level', 'Section',
           'Original Time Schedule', 'Parse Status', 'Notes']


def convert(src, program, out_path):
    doc = docx.Document(src)
    if os.path.exists(out_path):
        wb = load_workbook(out_path)
        ws = wb['Schedule']
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Schedule'
        ws.append(HEADERS)
        for c in ws[1]:
            c.font = Font(bold=True)

    n_rows = n_bad = 0
    ordinal = {1: '1st', 2: '2nd', 3: '3rd', 4: '4th', 5: '5th'}
    for tbl in doc.tables:
        year = None
        in_data = False
        for row in tbl.rows:
            cells = [c.text for c in row.cells]
            texts = [clean_cell(c) for c in cells]
            if len(set(texts)) == 1:               # merged banner row
                y = year_from_header(texts[0])
                if y:
                    year = y
                continue
            if texts[0].upper() == 'DAY':
                in_data = True
                continue
            if not in_data or texts[0].upper().startswith('TOTAL'):
                in_data = False if texts[0].upper().startswith('TOTAL') else in_data
                continue
            day_s, time_s, code, title, units, room, instr = texts[:7]
            if not code:
                continue
            code = CODE_FIXES.get(code, code)
            instr = NAME_FIXES.get(instr, instr)
            section = f'{program}-{year}A'
            base = dict(room=norm_room(room), instr=instr or 'TBA',
                        year=f'{ordinal.get(year, year)} Year', section=section)

            # A row can hold several meetings: the DAY and TIME cells pair up
            # line-by-line (e.g. 'Monday\nTuesday' with '1:00-2:00PM\n9:00-11:00AM').
            day_lines = [clean_cell(x) for x in cells[0].split('\n') if clean_cell(x)]
            time_lines = [clean_cell(x) for x in cells[1].split('\n') if clean_cell(x)]
            if len(day_lines) > 1 and len(day_lines) == len(time_lines):
                meetings = list(zip(day_lines, time_lines))
            else:
                meetings = [(day_s, time_s)]

            parsed = []
            for d_s, t_s in meetings:
                days = parse_day_string(d_s)
                rng = parse_time_range(t_s)
                if not days or not rng:
                    parsed = None
                    break
                parsed.append((days, rng))
            orig = ' ; '.join(f'{d} {t}' for d, t in meetings)
            if parsed is None:
                n_bad += 1
                ws.append([code, title, units, SEMESTER, '', '', '', '', '',
                           base['room'], base['instr'], program, base['year'],
                           section, orig, 'UNMATCHED: docx parse', ''])
                continue
            multi = len(parsed) > 1
            for k, (days, (start, end)) in enumerate(parsed, 1):
                ordered = [d for d in DAY_ORDER if d in days]
                day_label = '-'.join(DAY_LABEL[d] for d in ordered)
                ws.append([code, title, units, SEMESTER,
                           day_label, f'{fmt(start)} - {fmt(end)}',
                           fmt(start), fmt(end),
                           f'{k}/{len(parsed)}' if multi else '',
                           base['room'], base['instr'], program, base['year'],
                           section, orig, 'OK', ''])
                n_rows += 1

    wb.save(out_path)
    print(f'{os.path.basename(src)} [{program}] -> {out_path}: '
          f'{n_rows} rows written, {n_bad} unmatched')


if __name__ == '__main__':
    if len(sys.argv) < 3:
        sys.exit(__doc__)
    src = os.path.expanduser(sys.argv[1])
    program = sys.argv[2]
    out = sys.argv[3] if len(sys.argv) > 3 else \
        os.path.expanduser('~/Desktop/import_cleaned_docx.xlsx')
    convert(src, program, out)
