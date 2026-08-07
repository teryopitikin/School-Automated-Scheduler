"""
v3 — split the consolidated source workbook (Sheet1) Time Schedule into
uniform Day + Time, emitting ONE ROW PER MEETING so a class with a
different time on each day becomes separate rows. Read-only on the
source; writes a fresh workbook. No DB import.
"""
import datetime
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill

SRC = '/Users/macbok16-a/Downloads/Consolidated (3) - BSED UPDATED.xlsx'
OUT = '/Users/macbok16-a/Desktop/import_cleaned.xlsx'

# Manual day fixes for rows whose cell has a time but no weekday.
# Keyed by Excel row number; the day is prepended before parsing.
MANUAL_DAY = {}

# Other targeted data fixes, keyed by Excel row number.
MANUAL_SEMESTER = {}
MANUAL_ROOM = {364: 'Asynchronous'}                  # was blank
# Replace the raw Time Schedule text for rows with a data-entry error.
MANUAL_SCHEDULE = {
    # FS 1 / FS 2 (BSED-ENG-4A): Time Schedule blank in this revision;
    # Saturday slots carried over from the Consolidated_1 import.
    95: 'SAT 08:00 AM - 10:00 AM',
    96: 'SAT 10:00 AM - 12:00 NN',
    316: 'MWF 03:00 PM - 04:00 PM',  # CRIM 5 (3D-WHORL) source typo: was 03:00 AM
}
# Subject Code cell errors (copy-paste slips in the source sheet).
MANUAL_SUBJECT = {321: 'LEA 2'}   # cell literally contained '08:00 AM - 09:00 AM'

# ---------------------------------------------------------------- days
DAY_ORDER = ['MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT', 'SUN']
ABBREV = {'MON': 'M', 'TUE': 'T', 'WED': 'W', 'THU': 'TH',
          'FRI': 'F', 'SAT': 'Sat', 'SUN': 'Sun'}
FULLNAME = {
    'MONDAY': 'MON', 'MON': 'MON',
    'TUESDAY': 'TUE', 'TUES': 'TUE', 'TUE': 'TUE', 'TU': 'TUE',
    'WEDNESDAY': 'WED', 'WED': 'WED', 'WE': 'WED',
    'THURSDAY': 'THU', 'THURS': 'THU', 'THUR': 'THU', 'THU': 'THU', 'TH': 'THU',
    'FRIDAY': 'FRI', 'FRI': 'FRI', 'FR': 'FRI',
    'SATURDAY': 'SAT', 'SAT': 'SAT',
    'SUNDAY': 'SUN', 'SUN': 'SUN',
}


def _chartok(s):
    s = s.upper()
    out, i = [], 0
    while i < len(s):
        if s[i:i+3] in ('SUN', 'SAT'):
            out.append(s[i:i+3]); i += 3; continue
        if s[i:i+2] == 'TH':
            out.append('THU'); i += 2; continue
        if s[i:i+2] == 'SU':
            out.append('SUN'); i += 2; continue
        if s[i:i+2] == 'SA':
            out.append('SAT'); i += 2; continue
        one = {'M': 'MON', 'T': 'TUE', 'W': 'WED', 'F': 'FRI', 'S': 'SAT'}.get(s[i])
        if one is None:
            return None
        out.append(one); i += 1
    return out


def _map_token(tok):
    tok = tok.strip().upper().rstrip('.')
    if not tok:
        return []
    if tok in FULLNAME:
        return [FULLNAME[tok]]
    return _chartok(tok)


def parse_days(daystr):
    """List of day codes (unordered accumulation) or None if a token is unknown."""
    tokens = re.split(r'[\s,&/\-]+', daystr.strip())
    result = []
    for tok in tokens:
        mapped = _map_token(tok)
        if mapped is None:
            return None
        result.extend(mapped)
    return result or None


def order_days(codes):
    seen, out = set(), []
    for d in DAY_ORDER:
        if d in codes and d not in seen:
            seen.add(d); out.append(d)
    return out


def uniform_days(codes):
    return '-'.join(ABBREV[d] for d in order_days(codes))


# ---------------------------------------------------------------- times
def _school_hour(h):
    """Bare hour -> 24h in a school-day context (noon=12, 1-6 -> PM, 7-11 -> AM)."""
    if h == 12:
        return 12
    if 1 <= h <= 6:
        return h + 12
    return h


def _to24(h, m, mer):
    m = m or 0
    if mer == 'NN':
        return datetime.time(12, m)
    if mer == 'PM' and h != 12:
        h += 12
    elif mer == 'AM' and h == 12:
        h = 0
    return datetime.time(h % 24, m)


def parse_range(expr):
    expr = expr.replace('–', '-').replace('—', '-')
    toks = [t for t in re.findall(r'(\d{1,2})(?::(\d{2}))?\s*(AM|PM|NN)?', expr, re.I)
            if t[0] != '']
    if len(toks) < 2:
        return None
    (h1, m1, mer1), (h2, m2, mer2) = toks[0], toks[1]
    h1, h2 = int(h1), int(h2)
    m1 = int(m1) if m1 else 0
    m2 = int(m2) if m2 else 0
    mer1 = mer1.upper() if mer1 else None
    mer2 = mer2.upper() if mer2 else None
    inferred = False
    if mer1 is None and mer2 is not None:
        mer1 = ('AM' if h1 < 12 else 'PM') if mer2 == 'NN' else mer2
        inferred = True
    if mer2 is None and mer1 is not None:
        mer2 = mer1 if mer1 != 'NN' else 'PM'
    if mer1 is None and mer2 is None:
        s24, e24 = _school_hour(h1), _school_hour(h2)
        start, end = datetime.time(s24, m1), datetime.time(e24, m2)
        if end <= start and e24 < 12:      # end still looks earlier -> it's PM
            end = datetime.time(e24 + 12, m2)
    else:
        start, end = _to24(h1, m1, mer1), _to24(h2, m2, mer2)
        if inferred and mer1 == mer2 and start >= end:
            alt = _to24(h1, m1, 'AM' if mer1 == 'PM' else 'PM')
            if alt < end:
                start = alt
    # domain rule: classes never run at midnight; a 12:00 that became 00:00 is noon
    if start.hour == 0:
        start = datetime.time(12, start.minute)
    if end.hour == 0:
        end = datetime.time(12, end.minute)
    return start, end


RANGE_RE = re.compile(
    r'\d{1,2}(?::\d{2})?\s*(?:AM|PM|NN)?\s*[-–—]\s*\d{1,2}(?::\d{2})?\s*(?:AM|PM|NN)?',
    re.I)


def find_ranges(text):
    out = []
    for m in RANGE_RE.finditer(text):
        r = parse_range(m.group(0))
        if r:
            out.append(r)
    return out


def fmt(t):
    return t.strftime('%I:%M %p').lstrip('0')


# ---------------------------------------------------------------- meetings
def split_meetings(raw):
    """Return (meetings, status, reason). meeting = dict(days, start, end)."""
    s = str(raw).strip()
    if not s:
        return [], 'unmatched', 'empty'

    # 0) rotating-mode class: "<range> (<day> <week-rule>) <Mode>; ..." split on ';'
    if re.search(r'F2F|online|async', s, re.I) and ';' in s:
        meetings = []
        for seg in s.split(';'):
            seg = seg.strip()
            r = parse_range(seg)
            md = re.search(r'\(\s*([A-Za-z]+)([^)]*)\)', seg)
            mode = re.search(r'(F2F|online|async)', seg, re.I)
            if not (r and md):
                meetings = []
                break
            d = parse_days(md.group(1))
            if not d:
                meetings = []
                break
            bits = []
            if mode:
                bits.append(mode.group(1).upper() if mode.group(1).upper() == 'F2F'
                            else mode.group(1).capitalize())
            wk = md.group(2).strip()
            if wk:
                bits.append(wk)
            meetings.append({'days': d, 'start': r[0], 'end': r[1],
                             'note': ' · '.join(bits)})
        if meetings:
            return meetings, 'ok', ''
        return [], 'unmatched', 'online/async rotation'
    if re.search(r'online|async|F2F|every', s, re.I):
        return [], 'unmatched', 'online/async rotation'

    # 1) Friday-override:  <main> (F: <time>)
    mo = re.search(r'\(\s*F\s*:\s*([^)]*)\)', s, re.I)
    if mo:
        override = parse_range(mo.group(1))
        main = (s[:mo.start()] + s[mo.end():]).strip()
        meetings, st, _ = split_meetings(main)
        if st == 'unmatched':
            return [], 'unmatched', 'friday-override main unparsed'
        if override:
            newm = []
            for mtg in meetings:
                days = [d for d in mtg['days'] if d != 'FRI']
                if days:
                    newm.append({'days': days, 'start': mtg['start'], 'end': mtg['end']})
            newm.append({'days': ['FRI'], 'start': override[0], 'end': override[1]})
            meetings = newm
        return meetings, 'ok', ''

    # 2) per-range day markers:  <range> (T) <range> (Th)
    markers = re.findall(r'(\d[^()]*?(?:AM|PM|NN))\s*\(\s*([A-Za-z]{1,4})\s*\)', s, re.I)
    if len(markers) >= 2:
        meetings = []
        for rng_txt, day_txt in markers:
            d = parse_days(day_txt)
            r = parse_range(rng_txt)
            if d and r:
                meetings.append({'days': d, 'start': r[0], 'end': r[1]})
        if len(meetings) == len(markers):
            return meetings, 'ok', ''

    # 3) day-first: <day-groups> <ranges>
    ma = re.match(r'^([A-Za-z][A-Za-z&,/\.\- ]*?)\s+(\d.*)$', s)
    if ma:
        groups = [g for g in re.split(r'\s+', ma.group(1).strip()) if g]
        parsed = [parse_days(g) for g in groups]
        if all(parsed):
            ranges = find_ranges(ma.group(2))
            if not ranges:
                # no dash between the two times (e.g. "09:00 AM 10:00 AM")
                single = parse_range(ma.group(2))
                if single:
                    ranges = [single]
            if ranges:
                if len(parsed) == len(ranges) and len(ranges) > 1:
                    return ([{'days': parsed[i], 'start': ranges[i][0], 'end': ranges[i][1]}
                             for i in range(len(ranges))], 'ok', '')
                if len(parsed) == 1 and len(ranges) > 1:
                    d = parsed[0]
                    return ([{'days': d, 'start': r[0], 'end': r[1]} for r in ranges],
                            'ok', '')
                if len(ranges) == 1:
                    alldays = [d for g in parsed for d in g]
                    return ([{'days': alldays, 'start': ranges[0][0], 'end': ranges[0][1]}],
                            'ok', '')

    # 4) fallback single meeting — time-first (day at end) or last resort
    mb = re.match(r'^(.*?\d[^A-Za-z]*(?:AM|PM|NN)?)\s*-?\s*([A-Za-z][A-Za-z&,/\.\- ]*)$',
                  s, re.I)
    if mb and parse_days(mb.group(2)):
        r = parse_range(mb.group(1))
        if r:
            return [{'days': parse_days(mb.group(2)), 'start': r[0], 'end': r[1]}], 'ok', ''

    return [], 'unmatched', 'no day/time found'


# ---------------------------------------------------------------- run
wb = load_workbook(SRC, data_only=True)
ws = wb['Sheet1']
rows = list(ws.iter_rows(min_row=2, values_only=True))

out = Workbook()
o = out.active
o.title = 'Schedule'
headers = ['Subject Code', 'Descriptive Title', 'Units', 'Semester',
           'Day', 'Time', 'Time Start', 'Time End', 'Meeting',
           'Room Number', 'Faculty Name', 'Course', 'Year Level', 'Section',
           'Original Time Schedule', 'Parse Status', 'Notes']
o.append(headers)
for c in o[1]:
    c.font = Font(bold=True)

un = out.create_sheet('Unmatched')
un.append(['Excel Row', 'Subject Code', 'Section', 'Original Time Schedule', 'Reason'])
for c in un[1]:
    c.font = Font(bold=True)

split_fill = PatternFill('solid', fgColor='D6EAF8')   # blue = row was expanded
bad_fill = PatternFill('solid', fgColor='F8D7DA')     # red  = unmatched

n_rows_in = n_rows_out = n_split = n_bad = 0
unmatched = []

for idx, r in enumerate(rows, start=2):
    if r[0] in (None, ''):
        continue
    n_rows_in += 1
    base = [MANUAL_SUBJECT.get(idx, r[0]), r[1], r[2], MANUAL_SEMESTER.get(idx, r[3])]
    faculty = r[6] if (r[6] is not None and str(r[6]).strip() != '') else 'TBA'
    room = MANUAL_ROOM.get(idx, r[5])
    if room is not None:
        rs = str(room).strip()
        if re.fullmatch(r'\d+', rs):
            rs = f"Room {rs}"                     # number-only rooms -> "Room ##"
        else:
            rs = re.sub(r'(?i)^room\b', 'Room', rs)  # ROOM 15 -> Room 15
        room = rs
    tail = [room, faculty, r[7], r[8], r[9]]  # Room, Faculty, Course, Year, Section
    raw_e = MANUAL_SCHEDULE.get(idx, r[4])
    if idx in MANUAL_DAY:
        raw_e = f"{MANUAL_DAY[idx]} {raw_e}"
    meetings, status, reason = split_meetings(raw_e)

    if status == 'unmatched':
        n_bad += 1
        unmatched.append((idx, r[0], r[9], r[4], reason))
        o.append(base + ['', '', '', '', ''] + tail + [r[4], 'UNMATCHED: ' + reason, ''])
        for c in o[o.max_row]:
            c.fill = bad_fill
        un.append([idx, r[0], r[9], r[4], reason])
        continue

    multi = len(meetings) > 1
    if multi:
        n_split += 1
    ok_status = f"OK (day set: {MANUAL_DAY[idx]})" if idx in MANUAL_DAY else 'OK'
    for k, mtg in enumerate(meetings, 1):
        day = uniform_days(mtg['days'])
        tstr = f"{fmt(mtg['start'])} - {fmt(mtg['end'])}"
        part = f"{k}/{len(meetings)}" if multi else ''
        o.append(base + [day, tstr, fmt(mtg['start']), fmt(mtg['end']), part]
                 + tail + [r[4], ok_status, mtg.get('note', '')])
        n_rows_out += 1
        if multi:
            for c in o[o.max_row]:
                c.fill = split_fill

out.save(OUT)

# ---------------------------------------------------------------- report
print(f"Saved: {OUT}\n")
print(f"Source rows            : {n_rows_in}")
print(f"Output rows            : {n_rows_out + n_bad}")
print(f"  Source rows expanded : {n_split}  (different time per day -> multiple rows)")
print(f"  Unmatched            : {n_bad}")
print("\n--- UNMATCHED (need manual entry) ---")
for er, subj, sec, orig, reason in unmatched:
    print(f"  row {er:>3} | {reason:<22} | {subj} [{sec}] | {orig!r}")
