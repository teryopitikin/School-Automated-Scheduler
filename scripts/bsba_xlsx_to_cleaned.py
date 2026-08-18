"""Convert the BSBA subject-loading workbook (one sheet per section:
'HRM 1A', 'MM 2B', 'FM 4C', ...) into the cleaned-import format.

Each sheet has a header row ('Subject Code' / 'Descriptive Title' /
'Units' / 'Time Schedule' / 'Room Number' / 'Faculty Name') at a
sheet-specific column offset, and Time Schedule strings shaped
'<time range> - <DAYS>' (e.g. '4:00 -5:00 PM - MTTH', '8:00 9:00 PM-MWF',
'4:00_5:00 PM -FRI', '12:00 NN-3:00 PM- TUE').

Usage: python scripts/bsba_xlsx_to_cleaned.py <src.xlsx> [out.xlsx]
"""
import os
import re
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from docx_to_cleaned import (  # noqa: E402
    Emitter, HEADERS, clean_cell, norm_room, parse_day_string,
)

DAY_FIXES = {'MTTHI': 'MTTH'}   # sheet typo


def split_sched(raw):
    """'4:00 -5:00 PM - MTTH' -> ('4:00 -5:00 PM', 'MTTH') or None."""
    s = clean_cell(raw).replace('_', '-')
    m = re.match(r'^(.*?)[\s\-]*([A-Za-z]+)\.?$', s)
    if not m:
        return None
    time_s, day_tok = m.group(1), DAY_FIXES.get(m.group(2).upper(), m.group(2))
    if not parse_day_string(day_tok):
        return None
    return time_s, day_tok


def convert(src, out_path):
    wb_in = load_workbook(src, data_only=True)
    if os.path.exists(out_path):
        wb = load_workbook(out_path)
        ws = wb['Schedule']
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Schedule'
        ws.append(HEADERS)
        for c in ws[1]:
            c.font = Font(bold=True)
    em = Emitter(ws)

    for sheet in wb_in.sheetnames:
        m = re.match(r'^([A-Z]+)\s*(\d)\s*-?\s*([A-Z])$', sheet.strip())
        if not m:
            print(f'  ! skipping sheet {sheet!r} (name not <MAJOR> <year><letter>)')
            continue
        major, year, letter = m.group(1), int(m.group(2)), m.group(3)
        program = f'BSBA-{major}'
        section = f'{program} {year}{letter}'
        ws_in = wb_in[sheet]

        col = None
        for row in ws_in.iter_rows(values_only=False):
            texts = {clean_cell(str(c.value)) if c.value is not None else '': c.column - 1
                     for c in row}
            if 'Subject Code' in texts:
                col = {name: texts[name] for name in
                       ('Subject Code', 'Descriptive Title', 'Units',
                        'Time Schedule', 'Room Number', 'Faculty Name')
                       if name in texts}
                continue
            if col is None:
                continue
            cells = [c.value for c in row]
            def get(name):
                i = col.get(name)
                return clean_cell(str(cells[i])) if i is not None and i < len(cells) and cells[i] is not None else ''
            code = get('Subject Code')
            if not code or code == 'Subject Code':
                continue
            sched = get('Time Schedule')
            parts = split_sched(sched) if sched else None
            if parts is None:
                em.emit(code, get('Descriptive Title'), get('Units'),
                        norm_room(get('Room Number')), get('Faculty Name'),
                        program, year, section, [('', '')], sched or '(blank)')
                continue
            time_s, day_tok = parts
            em.emit(code, get('Descriptive Title'), get('Units'),
                    norm_room(get('Room Number')), get('Faculty Name'),
                    program, year, section, [(day_tok, time_s)],
                    sched)

    wb.save(out_path)
    print(f'{os.path.basename(src)} [BSBA] -> {out_path}: '
          f'{em.n_rows} rows written, {em.n_bad} unmatched')


if __name__ == '__main__':
    src = os.path.expanduser(sys.argv[1])
    out = sys.argv[2] if len(sys.argv) > 2 else \
        os.path.expanduser('~/Desktop/import_bsba.xlsx')
    convert(src, out)
