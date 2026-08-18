"""Convert the BSCRIM grid-timetable workbook ('REVISED CLASS SCHED & FAC
LOADING.xlsx', sheet 'CS 26-27') into the cleaned-import workbook format.

Layout: four year-level blocks side by side (TIME col + five section
columns each), a MWF row band and a TTH row band, and a merged SAT band
(CLJ 4 for all 3rd-year sections, CLJ 6 for all 4th-year sections).
Cells read 'COURSE, TEACHER[, day-override][, ROOM]'; vertically
contiguous identical cells are one class block. Day overrides like 'MW'
or 'MF' replace the band's days for that entry.

Usage: python scripts/grid_xlsx_to_cleaned.py <src.xlsx> [out.xlsx]
"""
import os
import re
import sys

from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from docx_to_cleaned import (  # noqa: E402
    Emitter, HEADERS, SEMESTER, clean_cell, parse_day_string,
    parse_time_range, norm_room, fmt, DAY_ORDER, DAY_LABEL, ORDINAL,
)
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font  # noqa: E402

PROGRAM = 'BSCRIM'
SHEET = 'CS 26-27'

# (year, time_col, [section_cols A..E])
BLOCKS = [(1, 1, [2, 3, 4, 5, 6]),
          (2, 7, [8, 9, 10, 11, 12]),
          (3, 13, [14, 15, 16, 17, 18]),
          (4, 19, [20, 21, 22, 23, 24])]
BANDS = [('MWF', range(5, 19)), ('TTH', range(22, 36))]
# Saturday review classes live in merged blocks; emit them directly.
SAT_CLASSES = [
    # (year, course, teacher, start-end)
    (3, 'CLJ 4', 'ATTY PULIDO', '1:00 PM - 6:00 PM'),
    (4, 'CLJ 6', 'ATTY CALAMBA', '1:00 PM - 5:00 PM'),
]

DAY_TOKEN = re.compile(r'^(?:M|T|W|TH|F|S|SAT|SUN)+$')


def parse_cell(text):
    """'CRIM 8, TEJANO, PRE REV, MF' -> (course, teacher, days_override, room)."""
    parts = [clean_cell(p) for p in text.split(',')]
    parts = [p for p in parts if p]
    course = parts[0]
    teacher = parts[1] if len(parts) > 1 else 'TBA'
    days = None
    room_parts = []
    for p in parts[2:]:
        if DAY_TOKEN.fullmatch(p.upper().replace(' ', '')) and parse_day_string(p):
            days = p
        elif p in ('- LEC', '-LEC', 'LEC'):
            continue                          # annotation noise
        else:
            room_parts.append(p)
    room = room_parts[-1] if room_parts else ('N/A' if teacher == 'TBA' else 'TBA')
    if room.upper() == 'CHEM':
        room = 'CHEMLAB'
    return course, teacher, days, norm_room(room)


def convert(src, out_path):
    wb_in = load_workbook(src, data_only=True)
    ws_in = wb_in[SHEET]

    if os.path.exists(out_path):
        wb = load_workbook(out_path)
        ws = wb['Schedule']
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Schedule'
        ws.append(HEADERS)
        for c in ws[1]:
            c.font = Font(bold=True)
    em = Emitter(ws)

    # section letters from the header row (col index -> 'A'..'E')
    letters = {}
    for _, _, sec_cols in BLOCKS:
        for i, ci in enumerate(sec_cols):
            letters[ci] = 'ABCDE'[i]

    for band_days, rows in BANDS:
        for year, time_col, sec_cols in BLOCKS:
            for ci in sec_cols:
                section = f'{PROGRAM} {year}{letters[ci]}'
                run = None   # (text, start, end)
                def flush(run):
                    if run is None:
                        return
                    text, start, end = run
                    course, teacher, dayq, room = parse_cell(text)
                    days = dayq or band_days
                    em.emit(course, '', '', room, teacher, PROGRAM, year,
                            section, [(days, f'{fmt(start)} - {fmt(end)}')],
                            f'{days} {fmt(start)}-{fmt(end)}')
                for ri in rows:
                    t_raw = ws_in.cell(row=ri, column=time_col).value
                    rng = parse_time_range(str(t_raw)) if t_raw else None
                    if rng is None:
                        flush(run); run = None
                        continue
                    start, end = rng
                    v = ws_in.cell(row=ri, column=ci).value
                    text = clean_cell(str(v)) if v is not None else ''
                    if not text:
                        flush(run); run = None
                        continue
                    if run and run[0] == text and run[2] == start:
                        run = (text, run[1], end)
                    else:
                        flush(run)
                        run = (text, start, end)
                flush(run)

    for year, course, teacher, rng_s in SAT_CLASSES:
        for letter in 'ABCDE':
            em.emit(course, '', '', 'TBA', teacher, PROGRAM, year,
                    f'{PROGRAM} {year}{letter}', [('SAT', rng_s)],
                    f'SAT {rng_s}')

    wb.save(out_path)
    print(f'{os.path.basename(src)} [{PROGRAM}] -> {out_path}: '
          f'{em.n_rows} rows written, {em.n_bad} unmatched')


if __name__ == '__main__':
    src = os.path.expanduser(sys.argv[1])
    out = sys.argv[2] if len(sys.argv) > 2 else \
        os.path.expanduser('~/Desktop/import_bscrim.xlsx')
    convert(src, out)
